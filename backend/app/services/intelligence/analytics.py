from __future__ import annotations

import csv
import io
from statistics import mean
from typing import Any

from openpyxl import load_workbook

from models.contracts import AnalyticsSummary


class AnalyticsService:
    async def summarize_text_dataset(self, dataset_name: str, content: str, kind: str) -> AnalyticsSummary:
        rows = self._rows_from_content(content, kind)
        return self._summarize_rows(dataset_name, rows)

    async def summarize_file(self, dataset_name: str, file_bytes: bytes, filename: str) -> AnalyticsSummary:
        lowered = filename.lower()
        if lowered.endswith(".xlsx"):
            rows = self._rows_from_excel(file_bytes)
        else:
            rows = self._rows_from_content(file_bytes.decode("utf-8", errors="ignore"), "csv")
        return self._summarize_rows(dataset_name, rows)

    def _rows_from_excel(self, file_bytes: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell) if cell is not None else f"column_{index}" for index, cell in enumerate(rows[0])]
        return [
            {headers[index]: value for index, value in enumerate(row)}
            for row in rows[1:]
        ]

    def _rows_from_content(self, content: str, kind: str) -> list[dict[str, Any]]:
        stripped = content.strip()
        if not stripped:
            return []
        if kind == "manual":
            lines = [line for line in stripped.splitlines() if line.strip()]
            return [{"entry": line} for line in lines]
        try:
            reader = csv.DictReader(io.StringIO(stripped))
            return list(reader)
        except csv.Error:
            return [{"entry": line} for line in stripped.splitlines() if line.strip()]

    def _summarize_rows(self, dataset_name: str, rows: list[dict[str, Any]]) -> AnalyticsSummary:
        if not rows:
            return AnalyticsSummary(
                dataset_name=dataset_name,
                row_count=0,
                column_count=0,
                insights=["No rows detected in the provided dataset."],
            )
        columns = list(rows[0].keys())
        numeric_columns: list[str] = []
        metrics: dict[str, Any] = {}
        for column in columns:
            values: list[float] = []
            for row in rows:
                value = row.get(column)
                try:
                    values.append(float(value))
                except (TypeError, ValueError):
                    continue
            if values:
                numeric_columns.append(column)
                metrics[column] = {
                    "min": min(values),
                    "max": max(values),
                    "mean": round(mean(values), 2),
                }
        insights = [
            f"Detected {len(rows)} rows across {len(columns)} columns.",
            f"Numeric columns: {', '.join(numeric_columns)}" if numeric_columns else "No numeric columns were detected.",
        ]
        return AnalyticsSummary(
            dataset_name=dataset_name,
            row_count=len(rows),
            column_count=len(columns),
            numeric_columns=numeric_columns,
            metrics=metrics,
            sample_rows=rows[:5],
            insights=insights,
        )

